
# -*- coding:utf-8 -*-
#模块功能：判断某个文件夹下有几个Excel文件，每个Excel有几个Sheet及Sheet Name
import os
import openpyxl
import xlrd
list = [] #所有数据

def getFileNames(path):
    filenames = os.listdir(path)
    for i, filename in enumerate(filenames):
         if i==0:
            iSpecialFile=i+1
            sFileName=filename

         print('==================第%s个文件========================='%(i+1))
         print('文件名：%s'%(filename))
         getSheetNames(path,filename)
    print('\n')
    print('--------------------选择指定的第几个文件-------------------------')
    print('指定的是第%s个文件:'%iSpecialFile+sFileName )
    print('----------------------------------------------------------------')

def getSheetNames(path,sFileName):
    temp = input("--------------------想读取第几列的数据：")
    col = int(temp) - 1
    wb = xlrd.open_workbook(path+'\\'+sFileName)
    # 获取workbook中所有的表格
    sheet = wb.sheet_by_index(0) # 根据sheet索引或者名称获取sheet内容 sheet索引从0开始
    print(sheet.name, sheet.nrows, sheet.ncols) # sheet的名称，行数，列数
    for row in sheet.col_values(col):  #列
        list.append(row)

def writeFile():
    f = openpyxl.Workbook()  # 创建工作簿
    # sheet1 = f.create_sheet()
    print("sheet names:", f.sheetnames)
    sheet1 = f.active   # 当前工作表的名称
    print("sheet1:", sheet1)
    row_file = len(list) # 生成5行
    col_three = 1 # 生成3列

    for row in range(row_file):
        for col in range(col_three):
            rw = row + 1
            cl = col + 1
            sheet1.cell(row=rw, column=cl, value=list[rw - 1])

    f.save("file/生成的Excel.xlsx")  # 保存文件

if __name__=='__main__':
    path=os.getcwd() + "\\" + r'Source'
    getFileNames(path)
    writeFile()
